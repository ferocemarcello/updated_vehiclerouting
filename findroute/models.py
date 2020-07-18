from django.db import models
from openpyxl import Workbook
from openpyxl.utils import column_index_from_string, coordinate_to_tuple, get_column_letter

import numpy
from geopy.distance import geodesic
import openpyxl
from sklearn.cluster import KMeans
import math

class XlsData:
    @classmethod
    def getXlsDataRange(cls,file, startCell, endCell, list_of_lists=False):
        wb = openpyxl.load_workbook(file)
        worksheet=wb.active
        rng = worksheet[startCell:endCell]#str:str
        data=[]
        if list_of_lists:
            for tuple in rng:
                row = ()
                for col in tuple:
                    row = row + (col.value,)
                data.append(list(row))
        else:
            for tuple in rng:
                row = ()
                for col in tuple:
                    row = row + (col.value,)
                data.append(row)
        return data

    @classmethod
    def getXlsDataStartOffset(cls,file, startCell, verticalOffset, horizontalOffset, list_of_lists=False):
        wb = openpyxl.load_workbook(file)
        worksheet=wb.active
        startCol = coordinate_to_tuple(startCell)[1]
        startRow = coordinate_to_tuple(startCell)[0]
        endRow= startRow + verticalOffset
        endCol=startCol+horizontalOffset
        endCell=get_column_letter(endCol)+str(endRow)
        rng = worksheet[startCell:endCell]#str:str
        data=[]
        if list_of_lists:
            for tuple in rng:
                row = ()
                for col in tuple:
                    row = row + (col.value,)
                data.append(list(row))
        else:
            for tuple in rng:
                row = ()
                for col in tuple:
                    row = row + (col.value,)
                data.append(row)
        return data

    @classmethod
    def writeMatrixToXls(cls,filename,matrix):#might be useful to write matrices to file
        wb=Workbook()
        sheet = wb.active
        for i in range(len(matrix)):
            for x in range(len(matrix[i])):
                sheet.cell(row=i+1, column=x+1).value = matrix[i][x]
        wb.save(filename)
    @classmethod
    def iterateRowTillFirstValueCell(cls, startCell, file, targetValue=None):
        startCol=coordinate_to_tuple(startCell)[1]
        startRow = coordinate_to_tuple(startCell)[0]
        wb = openpyxl.load_workbook(file)
        worksheet = wb.active
        row=[]
        #numcols=0
        for cell in worksheet.iter_cols(min_col=startCol, min_row=startRow, max_row=startRow):
            if cell[0].value == targetValue:
                break
            else:
                #numcols += 1
                row.append(tuple([float(i) for i in ((cell[0].value).split(')'))[1].split(';')]))
        '''lastcol=get_column_letter(startCol+numcols-1)
        lastcustomerdeparture=XlsData.getXlsDataRange(file, lastcol+str(startRow),lastcol+str(startRow), list_of_lists=False)
        lastcustomerdeparture = (lastcustomerdeparture[0])[0]
        lastcustomerdeparture =(lastcustomerdeparture.split(')'))[0]
        lastcustomerdeparture.replace("(", "")
        lastcustomerdeparture = (lastcustomerdeparture.split(';'))[1]
        row=(row,lastcustomerdeparture)'''
        return row

class DataInfo:
    @classmethod
    def computeRouteDistance(cls,listofcoords):
        totdist = 0
        for i in range(len(listofcoords)-1):
            totdist +=DataInfo.geodesic_distance_meters(listofcoords[i], listofcoords[i+1], as_integer=True)
        return totdist
    @classmethod
    def get_week_1_2_customers_indices(cls,all_data,num_weeks):
        one_week_frequence_customers_indices = [i for i in all_data.keys() if all_data[i]['frequence'] == 1]
        two_week_frequence_customers_indices = [i for i in all_data.keys() if
                                                all_data[i]['frequence'] == 2 and i != 0]  # excluding depot

        one_week_frequence_customers = [all_data[i]['gps_pos'] for i in all_data.keys() if
                                        all_data[i]['frequence'] == 1]
        week_clusterer = PositionClusterer(num_weeks, numpy.asarray([elem for elem in one_week_frequence_customers]),
                                           one_week_frequence_customers_indices)
        week_clusters = week_clusterer.compute_clustering()

        week1_customers_indices = two_week_frequence_customers_indices + week_clusters[0][1]
        week1_customers_indices.sort()
        week2_customers_indices = two_week_frequence_customers_indices + week_clusters[1][1]
        week2_customers_indices.sort()
        week_customer_indices={}
        week_customer_indices["week1"]=week1_customers_indices
        week_customer_indices["week2"] = week2_customers_indices
        return week_customer_indices

    @classmethod
    def get_day_clusters(cls, all_data, week1_customers_indices, week2_customers_indices):
        day_clusterer_week_1 = PositionClusterer(5, numpy.asarray(
            [all_data[i]['gps_pos'] for i in week1_customers_indices]),
                                                 week1_customers_indices)  # 5 is the number of days in a week
        day_clusterer_week_2 = PositionClusterer(5, numpy.asarray(
            [all_data[i]['gps_pos'] for i in week2_customers_indices]),
                                                 week2_customers_indices)  # 5 is the number of days in a week
        day_clusters_week_1 = day_clusterer_week_1.compute_clustering()
        day_clusters_week_2 = day_clusterer_week_2.compute_clustering()
        day_clusters = {}
        day_clusters[0] = day_clusters_week_1  # first week
        day_clusters[1] = day_clusters_week_2  # second week
        for i in day_clusters.keys():
            for j in day_clusters[i].keys():
                day_clusters[i][j][1].insert(0, 0)  # depot index
                day_clusters[i][j][1].sort()
        return day_clusters

    @classmethod
    def get_all_transportation_data(cls,pos_and_freq, tw_and_service_times, distances):
        all_data = {}
        for i in range(len(pos_and_freq)):
            posfreq = ((pos_and_freq[i][0], pos_and_freq[i][1]), pos_and_freq[i][2])
            twswk1 = (tw_and_service_times[i][0], tw_and_service_times[i][1], tw_and_service_times[i][2],
                      tw_and_service_times[i][3], tw_and_service_times[i][4])
            twewk1 = (tw_and_service_times[i][5], tw_and_service_times[i][6], tw_and_service_times[i][7],
                      tw_and_service_times[i][8], tw_and_service_times[i][9])
            twswk2 = (
                tw_and_service_times[i][10], tw_and_service_times[i][11], tw_and_service_times[i][12],
                tw_and_service_times[i][13],
                tw_and_service_times[i][14])
            twewk2 = (
                tw_and_service_times[i][15], tw_and_service_times[i][16], tw_and_service_times[i][17],
                tw_and_service_times[i][18],
                tw_and_service_times[i][19])
            sertime = tw_and_service_times[i][20]

            all_data[i] = {}
            all_data[i]['gps_pos'] = posfreq[0]
            all_data[i]['frequence'] = posfreq[1]
            all_data[i]['tw_wk1'] = [(twswk1[x], twewk1[x]) for x in range(len(twswk1))]
            all_data[i]['tw_wk2'] = [(twswk2[x], twewk2[x]) for x in range(len(twswk2))]
            all_data[i]['service_time'] = sertime
            all_data[i]['distances'] = distances[i]
        return all_data

    @classmethod
    def geodesic_distance_meters(cls,position_1, position_2, as_integer=False):#only works with gpsPositions
        """Computes the Manhattan distance between two points"""
        if as_integer:
            return int(geodesic(position_1,position_2).meters)
        else:
            return geodesic(position_1,position_2).meters

    @classmethod
    def geodesic_distance_matrix(cls,positions, as_integers=False):#may be useful to save the matrix then on file
        dist_mat=[]
        for  i in range(len(positions)):
            dist_mat.append([])
            for x in range(len(positions)):
                man = DataInfo.geodesic_distance_meters(positions[i], positions[x], as_integer=as_integers)
                dist_mat[i].insert(x,man)

        return dist_mat

    @classmethod
    def prepareRoutingData(cls, all_data, day_clusters, week, day):
        preproc={}
        preproc["coord"] = [all_data[z]['gps_pos'] for z in day_clusters[week][day][1]]
        tim = [None] * (len(day_clusters[week][day][1]))
        ser = [None] * (len(day_clusters[week][day][1]))
        if week == 0:
            for z in range(len(day_clusters[week][day][1])):
                tim[z] = (all_data[day_clusters[week][day][1][z]])['tw_wk1'][day]
        else:
            for z in range(len(day_clusters[week][day][1])):
                tim[z] = (all_data[day_clusters[week][day][1][z]])['tw_wk2'][day]
        preproc["time_windows"]=tim
        for z in range(len(day_clusters[week][day][1])):
            ser[z] = (all_data[day_clusters[week][day][1][z]])['service_time']
        preproc["service_times"]=ser
        dist = [None] * (len(day_clusters[week][day][1]))
        for z in range(len(day_clusters[week][day][1])):
            one_entry_dist = [None] * (len(day_clusters[week][day][1]))
            pos_ind = day_clusters[week][day][1][z]
            for x in range(len(day_clusters[week][day][1])):
                pos_ind2 = day_clusters[week][day][1][x]
                one_entry_dist[x] = all_data[pos_ind]['distances'][pos_ind2]
            dist[z] = one_entry_dist
        preproc["distances"]=dist
        preproc["global_indices"] = day_clusters[week][day][1]
        return preproc

    @classmethod
    def getPositionsFromIndices(cls, indices, all_data):
        return [all_data[i]['gps_pos'] for i in indices]

    @classmethod
    def getRoutesFromXls(cls, xlsfile,coords):
        depot=XlsData.getXlsDataRange(xlsfile, "A2", "A2", list_of_lists=False)
        depot=(depot[0])[0]
        depot.replace(" ","")
        depot = tuple([float(i) for i in depot.split(';')])
        route_coords_and_order = []
        for i in range(10):
            routeicustomerscoords = XlsData.iterateRowTillFirstValueCell(str("D"+str(i+3)), xlsfile, targetValue=None)
            routeicustomerscoords.insert(0, depot)
            routeicustomerscoords.append(depot)
            route_coords_and_order.append({})
            ord=[]
            dict={}
            for coord in routeicustomerscoords:
                ord.append(coords.index(coord))
            dict[0]=routeicustomerscoords
            dict[1]=ord
            route_coords_and_order[i]=dict
        return route_coords_and_order

    @classmethod
    def getRoutes(cls,xlsfile):
        depot = XlsData.getXlsDataRange(xlsfile, "A2", "A2", list_of_lists=False)
        depot = (depot[0])[0]
        depot.replace(" ", "")
        depot = tuple([float(i) for i in depot.split(';')])
        routeSummaries = []
        customercoords = []
        for i in range(10):
            departurei=XlsData.getXlsDataRange(xlsfile, str("B"+str(i+3)), str("B"+str(i+3)), list_of_lists=False)
            departurei = float((departurei[0])[0])
            arrivali = XlsData.getXlsDataRange(xlsfile, str("C" + str(i + 3)), str("C" + str(i + 3)),
                                                 list_of_lists=False)
            arrivali = float((arrivali[0])[0])
            routei = XlsData.iterateRowTillFirstValueCell(str("D"+str(i+3)), xlsfile, targetValue=None)

            customercoords = customercoords + [coord for coord in routei if coord not in customercoords]

            distanceRoute=DataInfo.computeRouteDistance([depot]+routei+[depot])
            routeDuration=arrivali-departurei
            routeSummary=[]
            routeSummary.append(distanceRoute)
            routeSummary.append(routeDuration)
            routeSummary.append(len(routei))
            routeSummaries.append(routeSummary)

        customercoords.insert(0, depot)
        return routeSummaries,customercoords

    @classmethod
    def getWeekDayNumber(cls, working_days_per_week,num_of_weeks, i):

        weeknumber = int(math.ceil((i + 1) / working_days_per_week))
        daynumber = int(i + 1 - (num_of_weeks * working_days_per_week / num_of_weeks) * (
            math.floor(i / working_days_per_week)))
        return weeknumber,daynumber


class PositionClusterer:
    def __init__(self, k,positions,position_indices):
        self.k=k#number of clusters
        self.positions=positions
        self.position_inidices=position_indices#len=len(positions)
    def compute_clustering(self):
        kmeans = KMeans(n_clusters=self.k).fit(self.positions)

        def getPositionsFromLabel(label,KMeansLabels,positions):
            return [(positions[i][0],positions[i][1]) for i in range(len(positions)) if KMeansLabels[i]==label]

        def getIndicesFromLabel(label,KMeansLabels,position_indices):
            return [position_indices[i] for i in range(len(KMeansLabels)) if KMeansLabels[i]==label]

        def aggregatePositionClusters(k):
            PositionClusters={}
            for i in range(k):
                PositionClusters[i]=(getPositionsFromLabel(i,kmeans.labels_,self.positions),getIndicesFromLabel(i,kmeans.labels_,self.position_inidices))
            return PositionClusters

        return aggregatePositionClusters(self.k)